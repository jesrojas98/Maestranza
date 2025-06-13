from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Sum, F, Count
from django.db import models
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.core.paginator import Paginator
from django.views.decorators.http import require_POST
from datetime import datetime, timedelta
from .models import *
from .forms import *
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from django.contrib.auth.models import User, Group
from django.db import transaction


# Importar decoradores de permisos
from .decorators import (
    solo_administrador, 
    gestor_o_admin, 
    todos_los_roles,
    puede_gestionar_productos,
    puede_gestionar_inventario,
    puede_asignar_ubicaciones,
    puede_ver_alertas,
    verificar_permiso_ajax,
    obtener_permisos_usuario
)

# Importar para autenticación
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import AuthenticationForm

# =================
# AUTENTICACIÓN
# =================

def login_view(request):
    """Vista de login personalizada"""
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            
            if user is not None:
                # Verificar si el usuario tiene perfil y está activo
                if hasattr(user, 'perfil') and user.perfil.activo:
                    login(request, user)
                    messages.success(request, f'Bienvenido, {user.get_full_name() or user.username}!')
                    
                    # Configurar sucursal inicial si tiene asignada
                    if user.perfil.sucursal_asignada:
                        request.session['sucursal_id'] = user.perfil.sucursal_asignada.id
                        request.session['sucursal_nombre'] = user.perfil.sucursal_asignada.nombre
                    
                    # Redireccionar según el parámetro 'next' o al dashboard
                    next_url = request.GET.get('next', 'dashboard')
                    return redirect(next_url)
                else:
                    messages.error(request, 'Su cuenta está inactiva. Contacte al administrador.')
            else:
                messages.error(request, 'Usuario o contraseña incorrectos.')
        else:
            messages.error(request, 'Por favor corrija los errores en el formulario.')
    else:
        form = AuthenticationForm()
    
    context = {
        'form': form,
        'title': 'Iniciar Sesión - Sistema de Inventario'
    }
    
    return render(request, 'registration/login.html', context)

@login_required
def logout_view(request):
    """Vista de logout"""
    username = request.user.get_full_name() or request.user.username
    logout(request)
    messages.success(request, f'Sesión cerrada correctamente. ¡Hasta luego, {username}!')
    return redirect('login')

# =================
# DASHBOARD
# =================
def dashboard(request):
    """Dashboard con estadísticas correctas sin duplicados"""
    
    # Manejar cambio de sucursal
    sucursal_param = request.GET.get('sucursal')
    if sucursal_param:
        try:
            sucursal_id = int(sucursal_param)
            sucursal = get_object_or_404(Sucursal, id=sucursal_id, activa=True)
            request.session['sucursal_id'] = sucursal_id
            request.session['sucursal_nombre'] = sucursal.nombre
            request.session.modified = True
        except (ValueError, TypeError):
            pass
    
    # Obtener sucursales y sucursal actual
    sucursales = Sucursal.objects.filter(activa=True)
    sucursal_actual = None
    sucursal_id = request.session.get('sucursal_id')
    if not sucursal_id:
        sucursales_first = Sucursal.objects.filter(activa=True).first()
        if sucursales_first:
            request.session['sucursal_id'] = sucursales_first.id
            request.session['sucursal_nombre'] = sucursales_first.nombre
            sucursal_id = sucursales_first.id
    
    try:
        sucursal_activa = Sucursal.objects.get(id=sucursal_id, activa=True)
    except Sucursal.DoesNotExist:
        sucursal_activa = Sucursal.objects.filter(activa=True).first()
        if sucursal_activa:
            request.session['sucursal_id'] = sucursal_activa.id
            request.session['sucursal_nombre'] = sucursal_activa.nombre
    
    if sucursal_id:
        try:
            sucursal_actual = Sucursal.objects.get(id=sucursal_id, activa=True)
        except Sucursal.DoesNotExist:
            sucursal_actual = None
    
    if not sucursal_actual and sucursales.exists():
        sucursal_actual = sucursales.first()
        request.session['sucursal_id'] = sucursal_actual.id
        request.session['sucursal_nombre'] = sucursal_actual.nombre
        request.session.modified = True
    
    # ESTADÍSTICAS CORRECTAS SIN DUPLICADOS
    if sucursal_actual:
        # 1. Total de productos en la sucursal (a través de inventario)
        total_productos = Inventario.objects.filter(
            sucursal=sucursal_actual
        ).values('producto').distinct().count()
        
        # 2. Productos con stock bajo (cantidad <= stock_minimo)
        productos_stock_bajo = Inventario.objects.filter(
            sucursal=sucursal_actual,
            cantidad__lte=F('producto__stock_minimo')
        ).values('producto').distinct().count()
        
        # 3. Alertas activas en esta sucursal
        alertas_stock = AlertaStock.objects.filter(
            status='activa',
            sucursal=sucursal_actual  # Solo si tienes campo sucursal
        ).values('producto').distinct().count()
        
        # 4. Productos por vencer (lotes con vencimiento próximo)
        fecha_limite = timezone.now().date() + timedelta(days=30)
        
        productos_vencer = Lote.objects.filter(
            producto__inventario__sucursal=sucursal_actual,
            fecha_vencimiento__lte=fecha_limite,
            fecha_vencimiento__gte=timezone.now().date()
        ).values('producto').distinct().count()
        
        # 5. Movimientos recientes de esta sucursal
        movimientos_recientes = []
        user_perfil = request.user.perfil    
        if user_perfil.puede_gestionar_inventario():
            movimientos_recientes = MovimientoInventario.objects.filter(
                sucursal=sucursal_activa
            ).select_related('producto', 'usuario').order_by('-fecha')[:5]
        
    else:
        # Sin sucursal seleccionada - estadísticas globales
        total_productos = Producto.objects.filter(activo=True).count()
        productos_stock_bajo = 0
        alertas_stock = AlertaStock.objects.filter(status='activa').count()
        productos_vencer = 0
        movimientos_recientes = MovimientoInventario.objects.select_related(
            'producto', 'usuario'
        ).order_by('-fecha')[:10]

    user_perfil = request.user.perfil    
    if user_perfil.puede_ver_alertas():
            # Gestores e admins ven alertas activas
            alertas_count = AlertaStock.objects.filter(
                status='activa',
                sucursal=sucursal_activa
            ).count()
    else:
            alertas_count = 0
        
    stats = {
        'total_productos': Producto.objects.filter(activo=True).count(),
        'total_stock': Inventario.objects.filter(
            sucursal=sucursal_activa
        ).aggregate(
            total=models.Sum('cantidad')
        )['total'] or 0,
        'productos_bajo_stock': Inventario.objects.filter(
            sucursal=sucursal_activa,
            cantidad__lte=models.F('producto__stock_minimo')
        ).count(),
        'alertas_activas': alertas_count,
    }
    
    context = {
        'sucursales': sucursales,
        'sucursal_actual': sucursal_actual,
        'total_productos': total_productos,
        'productos_stock_bajo': productos_stock_bajo,
        'alertas_stock': alertas_stock,
        'productos_vencer': productos_vencer,
        'movimientos_recientes': movimientos_recientes,
    }

    
    return render(request, 'inventario/dashboard.html', context)

@todos_los_roles
def lista_productos(request):
    """Vista para listar productos con filtros, búsqueda y estadísticas"""
    
    # Obtener parámetros de filtrado (mantenemos tu lógica existente)
    busqueda = request.GET.get('q', '').strip()
    categoria_seleccionada = request.GET.get('categoria', '')
    estado = request.GET.get('estado', '')  # Nuevo filtro
    
    # Query base
    productos = Producto.objects.select_related('categoria')
    
    # Aplicar filtros de búsqueda (tu lógica existente)
    if busqueda:
        productos = productos.filter(
            Q(codigo__icontains=busqueda) | 
            Q(nombre__icontains=busqueda) |
            Q(descripcion__icontains=busqueda)
        )
    
    if categoria_seleccionada:
        productos = productos.filter(categoria_id=categoria_seleccionada)
    
    # Nuevo filtro por estado
    if estado == 'activo':
        productos = productos.filter(activo=True)
    elif estado == 'inactivo':
        productos = productos.filter(activo=False)
    
    # Ordenamiento
    productos = productos.order_by('nombre')
    
    # Estadísticas para el nuevo panel
    total_productos = productos.count()
    productos_activos = productos.filter(activo=True).count()
    productos_con_imagen = productos.exclude(imagen='').count()
    
    # CORRECCIÓN: Obtener categorías para filtros
    # Cambiar filter(activa=True) por all() si da error, o verificar el nombre del campo
    try:
        categorias = Categoria.objects.filter(activa=True).annotate(
            productos_count=Count('producto')
        ).order_by('nombre')
    except:
        # Si falla, usar todas las categorías
        categorias = Categoria.objects.annotate(
            productos_count=Count('producto')
        ).order_by('nombre')
    
    # Paginación (mantenemos tu configuración)
    paginator = Paginator(productos, 12)  # 12 productos por página para cards
    page_number = request.GET.get('page')
    productos_page = paginator.get_page(page_number)
    
    context = {
        'productos': productos_page,
        'categorias': categorias,
        'busqueda': busqueda,
        'categoria_seleccionada': int(categoria_seleccionada) if categoria_seleccionada else None,
        'estado': estado,
        
        # Nuevas estadísticas para el template
        'productos_activos': productos_activos,
        'productos_con_imagen': productos_con_imagen,
    }
    
    return render(request, 'inventario/productos/lista.html', context)

@todos_los_roles
def detalle_producto(request, producto_id):
    """Vista para ver detalle completo de un producto con imagen"""
    print("🔍 INICIANDO vista detalle_producto")
    
    try:
        producto = get_object_or_404(Producto, id=producto_id)
        print(f"✅ Producto obtenido: {producto}")
    except Exception as e:
        print(f"❌ Error obteniendo producto: {e}")
        raise
    
    try:
        # Obtener inventario por sucursal
        inventarios = Inventario.objects.filter(
            producto=producto
        ).select_related('sucursal', 'ubicacion_detallada').order_by('sucursal__nombre')
        print(f"✅ Inventarios obtenidos: {inventarios.count()}")
    except Exception as e:
        print(f"❌ Error obteniendo inventarios: {e}")
        inventarios = []
    
    try:
        # Calcular stock total
        stock_total = sum(inv.cantidad for inv in inventarios)
        print(f"✅ Stock total calculado: {stock_total}")
    except Exception as e:
        print(f"❌ Error calculando stock: {e}")
        stock_total = 0
    
    try:
        # Obtener sucursal activa de la sesión
        sucursal_id = request.session.get('sucursal_id')
        sucursal_activa = None
        if sucursal_id:
            try:
                sucursal_activa = Sucursal.objects.get(id=sucursal_id)
            except Sucursal.DoesNotExist:
                sucursal_activa = Sucursal.objects.filter(activa=True).first()
                if sucursal_activa:
                    request.session['sucursal_id'] = sucursal_activa.id
                    request.session['sucursal_nombre'] = sucursal_activa.nombre
        
        if not sucursal_activa:
            sucursal_activa = Sucursal.objects.filter(activa=True).first()
        print(f"✅ Sucursal activa: {sucursal_activa}")
    except Exception as e:
        print(f"❌ Error obteniendo sucursal: {e}")
        sucursal_activa = None
    
    try:
        # Obtener movimientos recientes
        movimientos = MovimientoInventario.objects.filter(
            producto=producto
        ).select_related('sucursal', 'usuario').order_by('-fecha')[:10]
        print(f"✅ Movimientos obtenidos: {movimientos.count()}")
    except Exception as e:
        print(f"❌ Error obteniendo movimientos: {e}")
        movimientos = []
    
    try:
        # Obtener alertas activas
        alertas_activas = AlertaStock.objects.filter(
            producto=producto,
            status='activa'
        ).select_related('sucursal').order_by('-fecha_creacion')
        print(f"✅ Alertas obtenidas: {alertas_activas.count()}")
    except Exception as e:
        print(f"❌ Error obteniendo alertas: {e}")
        alertas_activas = []
    
    # ✅ HISTORIAL DE PRECIOS CON DEBUG EXTREMO
    historial_precios = []
    historial_precios_producto = []
    historial_precios_proveedores = []
    
    print("🔍 INICIANDO obtención de historial de precios...")
    
    try:
        print("🔍 Intentando importar HistorialPrecios...")
        from .models import HistorialPrecios
        print("✅ HistorialPrecios importado correctamente")
        
        # ✅ QUERY MÁS BÁSICA POSIBLE
        print("🔍 Ejecutando query básica de HistorialPrecios...")
        historial_simple = HistorialPrecios.objects.filter(producto=producto)
        print(f"✅ Query básica ejecutada, count: {historial_simple.count()}")
        
        print("🔍 Intentando obtener campos específicos...")
        # Probar query simple sin select_related
        historial_precios = list(historial_simple.order_by('-fecha')[:10])
        print(f"✅ Historial básico obtenido: {len(historial_precios)} registros")
        
        print("🔍 Separando por proveedor...")
        historial_precios_producto = [h for h in historial_precios if h.proveedor is None]
        historial_precios_proveedores = [h for h in historial_precios if h.proveedor is not None]
        print(f"✅ Separación completada - Producto: {len(historial_precios_producto)}, Proveedores: {len(historial_precios_proveedores)}")
        
    except ImportError as e:
        print(f"❌ Error importando HistorialPrecios: {e}")
        historial_precios = []
        historial_precios_producto = []
        historial_precios_proveedores = []
    except Exception as e:
        print(f"❌ ERROR EN HISTORIAL DE PRECIOS: {e}")
        print(f"❌ Tipo de error: {type(e)}")
        import traceback
        traceback.print_exc()
        # ✅ CONTINUAR SIN HISTORIAL
        historial_precios = []
        historial_precios_producto = []
        historial_precios_proveedores = []
    
    try:
        # Obtener lotes si el producto los maneja
        lotes = []
        if producto.maneja_lotes:
            try:
                from .models import Lote
                lotes = Lote.objects.filter(
                    producto=producto
                ).select_related('proveedor').order_by('-fecha_compra')[:5]
                print(f"✅ Lotes obtenidos: {len(lotes)}")
            except ImportError:
                print("⚠️ Modelo Lote no encontrado")
                pass
    except Exception as e:
        print(f"❌ Error obteniendo lotes: {e}")
        lotes = []
    
    try:
        # ✅ ESTADÍSTICAS BÁSICAS
        estadisticas_precio = {
            'tiene_precio': producto.precio is not None and producto.precio > 0,
            'precio_formateado': f"${producto.precio:,.2f}" if producto.precio else "Sin precio",
            'necesita_actualizacion': False,
            'fecha_ultima_actualizacion': producto.precio_actualizado,
        }
        print(f"✅ Estadísticas de precio creadas: {estadisticas_precio}")
    except Exception as e:
        print(f"❌ Error creando estadísticas de precio: {e}")
        estadisticas_precio = {'tiene_precio': False, 'precio_formateado': 'Error'}
    
    try:
        # ✅ ÚLTIMO CAMBIO SIMPLIFICADO
        ultimo_cambio_precio = None
        if historial_precios_producto:
            ultimo = historial_precios_producto[0]
            ultimo_cambio_precio = {
                'usuario': ultimo.usuario,
                'fecha_cambio': ultimo.fecha,
                'precio_actual': ultimo.precio
            }
        print(f"✅ Último cambio de precio: {ultimo_cambio_precio}")
    except Exception as e:
        print(f"❌ Error obteniendo último cambio: {e}")
        ultimo_cambio_precio = None
    
    try:
        # ✅ VARIACIÓN SIMPLE
        variacion_precio = None
        if len(historial_precios_producto) >= 2:
            precio_actual = historial_precios_producto[0].precio
            precio_anterior = historial_precios_producto[1].precio
            if precio_anterior and precio_anterior > 0:
                variacion = ((precio_actual - precio_anterior) / precio_anterior) * 100
                variacion_precio = {
                    'porcentaje': round(variacion, 2),
                    'tipo': 'aumento' if variacion > 0 else 'disminucion' if variacion < 0 else 'sin_cambio',
                    'precio_anterior': precio_anterior,
                    'precio_actual': precio_actual
                }
        print(f"✅ Variación de precio calculada: {variacion_precio}")
    except Exception as e:
        print(f"❌ Error calculando variación: {e}")
        variacion_precio = None
    
    print("🔍 Preparando context...")
    context = {
        'producto': producto,
        'inventarios': inventarios,
        'stock_total': stock_total,
        'sucursal_activa': sucursal_activa,
        'movimientos': movimientos,
        'alertas_activas': alertas_activas,
        'historial_precios': historial_precios,
        'historial_precios_producto': historial_precios_producto,
        'historial_precios_proveedores': historial_precios_proveedores,
        'ultimo_cambio_precio': ultimo_cambio_precio,
        'estadisticas_precio': estadisticas_precio,
        'variacion_precio': variacion_precio,
        'lotes': lotes,
        'necesita_reposicion': stock_total <= producto.stock_minimo
    }
    
    print("✅ Context preparado, renderizando template...")
    return render(request, 'inventario/productos/detalle.html', context)

@puede_gestionar_productos
def crear_producto(request):
    """Vista para crear un nuevo producto con imagen"""
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Datos básicos del producto
                codigo = request.POST.get('codigo', '').strip().upper()
                nombre = request.POST.get('nombre', '').strip()
                descripcion = request.POST.get('descripcion', '').strip()
                categoria_id = request.POST.get('categoria')
                unidad_medida = request.POST.get('unidad_medida', 'unidad')
                stock_minimo = int(request.POST.get('stock_minimo', 0))
                maneja_lotes = request.POST.get('maneja_lotes') == 'on'
                maneja_vencimiento = request.POST.get('maneja_vencimiento') == 'on'
                activo = request.POST.get('activo', 'on') == 'on'  # Por defecto activo
                
                # Validaciones básicas
                if not codigo or not nombre:
                    messages.error(request, 'El código y nombre son obligatorios.')
                    return render(request, 'inventario/productos/crear.html', {
                        'categorias': Categoria.objects.all().order_by('nombre'),
                        'etiquetas': Etiqueta.objects.all(),
                        'form_data': request.POST
                    })
                
                # Verificar código único
                if Producto.objects.filter(codigo=codigo).exists():
                    messages.error(request, f'Ya existe un producto con el código "{codigo}".')
                    return render(request, 'inventario/productos/crear.html', {
                        'categorias': Categoria.objects.all().order_by('nombre'),
                        'etiquetas': Etiqueta.objects.all(),
                        'form_data': request.POST
                    })
                
                # Obtener categoría
                categoria = get_object_or_404(Categoria, id=categoria_id) if categoria_id else None
                
                # Crear producto
                producto = Producto.objects.create(
                    codigo=codigo,
                    nombre=nombre,
                    descripcion=descripcion,
                    categoria=categoria,
                    unidad_medida=unidad_medida,
                    stock_minimo=stock_minimo,
                    maneja_lotes=maneja_lotes,
                    maneja_vencimiento=maneja_vencimiento,
                    activo=activo
                )
                
                # Manejar imagen si se subió
                imagen = request.FILES.get('imagen')
                if imagen:
                    # Validar tipo de archivo
                    ext = os.path.splitext(imagen.name)[1].lower()
                    if ext not in ['.jpg', '.jpeg', '.png', '.gif', '.webp']:
                        messages.error(request, 'Formato de imagen no válido. Use JPG, PNG, GIF o WebP.')
                        producto.delete()
                        return render(request, 'inventario/productos/crear.html', {
                            'categorias': Categoria.objects.all().order_by('nombre'),
                            'etiquetas': Etiqueta.objects.all(),
                            'form_data': request.POST
                        })
                    
                    # Validar tamaño (5MB máximo)
                    if imagen.size > 5 * 1024 * 1024:
                        messages.error(request, 'La imagen es demasiado grande. Máximo 5MB.')
                        producto.delete()
                        return render(request, 'inventario/productos/crear.html', {
                            'categorias': Categoria.objects.all().order_by('nombre'),
                            'etiquetas': Etiqueta.objects.all(),
                            'form_data': request.POST
                        })
                    
                    # Guardar imagen
                    producto.imagen = imagen
                    producto.save()
                
                # Manejar etiquetas
                etiquetas_ids = request.POST.getlist('etiquetas')
                if etiquetas_ids:
                    etiquetas = Etiqueta.objects.filter(id__in=etiquetas_ids)
                    producto.etiquetas.set(etiquetas)
                
                # Crear inventario inicial en todas las sucursales
                sucursales = Sucursal.objects.filter(activa=True)
                for sucursal in sucursales:
                    Inventario.objects.get_or_create(
                        producto=producto,
                        sucursal=sucursal,
                        defaults={'cantidad': 0}
                    )
                
                messages.success(request, f'Producto "{producto.nombre}" creado exitosamente.')
                return redirect('detalle_producto', producto_id=producto.id)
                
        except ValueError as e:
            messages.error(request, f'Error en los datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Error al crear el producto: {str(e)}')
    
    # GET request o error en POST
    context = {
        'categorias': Categoria.objects.all().order_by('nombre'),
        'etiquetas': Etiqueta.objects.all(),
        'form_data': request.POST if request.method == 'POST' else {}
    }
    return render(request, 'inventario/productos/crear.html', context)

@puede_gestionar_inventario
def movimiento_inventario(request, producto_id):
    """Registrar movimiento de inventario"""
    producto = get_object_or_404(Producto, id=producto_id)
    
    if request.method == 'POST':
        form = MovimientoForm(request.POST, producto=producto)
        if form.is_valid():
            movimiento = form.save(commit=False)
            movimiento.producto = producto
            movimiento.usuario = request.user
            
            # Obtener inventario actual
            inventario, created = Inventario.objects.get_or_create(
                producto=producto,
                sucursal=movimiento.sucursal,
                defaults={'cantidad': 0}
            )
            
            movimiento.cantidad_anterior = inventario.cantidad
            
            # Calcular nueva cantidad según el tipo de movimiento
            if movimiento.tipo == 'entrada':
                nueva_cantidad = inventario.cantidad + movimiento.cantidad
            elif movimiento.tipo == 'salida':
                nueva_cantidad = inventario.cantidad - movimiento.cantidad
                if nueva_cantidad < 0:
                    messages.error(request, 'No hay suficiente stock para realizar la salida.')
                    return render(request, 'inventario/movimientos/crear.html', {
                        'form': form, 
                        'producto': producto
                    })
            elif movimiento.tipo == 'ajuste':
                nueva_cantidad = movimiento.cantidad
                movimiento.cantidad = nueva_cantidad - inventario.cantidad
            elif movimiento.tipo == 'transferencia':
                nueva_cantidad = inventario.cantidad - movimiento.cantidad
                if nueva_cantidad < 0:
                    messages.error(request, 'No hay suficiente stock para la transferencia.')
                    return render(request, 'inventario/movimientos/crear.html', {
                        'form': form, 
                        'producto': producto
                    })
                
                # Actualizar inventario destino
                inventario_destino, _ = Inventario.objects.get_or_create(
                    producto=producto,
                    sucursal=movimiento.sucursal_destino,
                    defaults={'cantidad': 0}
                )
                inventario_destino.cantidad += movimiento.cantidad
                inventario_destino.save()
            
            movimiento.cantidad_nueva = nueva_cantidad
            movimiento.save()
            
            # Actualizar inventario origen
            inventario.cantidad = nueva_cantidad
            inventario.save()
            
            # Verificar si necesita crear alerta de stock bajo
            if inventario.cantidad <= producto.stock_minimo:
                AlertaStock.objects.get_or_create(
                    tipo='stock_bajo',
                    producto=producto,
                    sucursal=movimiento.sucursal,
                    defaults={
                        'mensaje': f'Stock bajo para {producto.nombre} en {movimiento.sucursal.nombre}. '
                                 f'Stock actual: {inventario.cantidad}, Mínimo: {producto.stock_minimo}'
                    }
                )
            
            messages.success(request, 'Movimiento registrado exitosamente.')
            return redirect('detalle_producto', producto_id=producto.id)
    else:
        form = MovimientoForm(producto=producto)
    
    return render(request, 'inventario/movimientos/crear.html', {
        'form': form, 
        'producto': producto
    })

@puede_ver_alertas
def alertas_stock(request):
    """Lista de alertas de stock"""
    alertas = AlertaStock.objects.filter(
        status='activa'
    ).select_related('producto', 'sucursal', 'lote').order_by('-fecha_creacion')
    
    # Filtros
    tipo_filter = request.GET.get('tipo')
    sucursal_filter = request.GET.get('sucursal')
    
    if tipo_filter:
        alertas = alertas.filter(tipo=tipo_filter)
    
    if sucursal_filter:
        alertas = alertas.filter(sucursal_id=sucursal_filter)
    
    # Calcular estadísticas ANTES de la paginación
    total_alertas = alertas.count()
    alertas_stock_bajo = alertas.filter(tipo='stock_bajo').count()
    alertas_vencimiento = alertas.filter(tipo='vencimiento').count()
    
    # Paginación
    paginator = Paginator(alertas, 20)
    page_number = request.GET.get('page')
    alertas_paginadas = paginator.get_page(page_number)
    
    context = {
        'alertas': alertas_paginadas,
        'sucursales': Sucursal.objects.filter(activa=True),
        'tipos_alerta': AlertaStock.TIPO_CHOICES,
        'tipo_seleccionado': tipo_filter,
        'sucursal_seleccionada': sucursal_filter,
        # Estadísticas
        'total_alertas': total_alertas,
        'alertas_stock_bajo': alertas_stock_bajo,
        'alertas_vencimiento': alertas_vencimiento,
    }
    
    return render(request, 'inventario/alertas/lista.html', context)

@puede_ver_alertas
def resolver_alerta(request, alerta_id):
    """Resolver una alerta de stock"""
    alerta = get_object_or_404(AlertaStock, id=alerta_id)
    
    if request.method == 'POST':
        alerta.status = 'resuelta'
        alerta.fecha_resolucion = timezone.now()
        alerta.usuario_resolucion = request.user
        alerta.save()
        
        messages.success(request, 'Alerta marcada como resuelta.')
    
    return redirect('alertas_stock')

@todos_los_roles
def lista_proveedores(request):
    """Lista de proveedores"""
    proveedores = Proveedor.objects.filter(activo=True)
    
    busqueda = request.GET.get('q')
    if busqueda:
        proveedores = proveedores.filter(
            Q(nombre__icontains=busqueda) |
            Q(contacto__icontains=busqueda)
        )
    
    # Calcular estadísticas ANTES de la paginación
    total_proveedores = proveedores.count()
    proveedores_activos = proveedores.filter(activo=True).count()  # Redundante si ya filtras por activo=True
    proveedores_con_email = proveedores.exclude(email__isnull=True).exclude(email__exact='').count()
    proveedores_con_telefono = proveedores.exclude(telefono__isnull=True).exclude(telefono__exact='').count()
    
    paginator = Paginator(proveedores, 20)
    page_number = request.GET.get('page')
    proveedores_paginados = paginator.get_page(page_number)
    
    context = {
        'proveedores': proveedores_paginados,
        'busqueda': busqueda or '',
        # Estadísticas
        'total_proveedores': total_proveedores,
        'proveedores_activos': proveedores_activos,
        'proveedores_con_email': proveedores_con_email,
        'proveedores_con_telefono': proveedores_con_telefono,
    }
    
    return render(request, 'inventario/proveedores/lista.html', context)
    
@puede_gestionar_productos
def editar_producto(request, producto_id):
    """Vista para editar un producto existente con imagen"""
    producto = get_object_or_404(Producto, id=producto_id)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Datos básicos del producto
                codigo = request.POST.get('codigo', '').strip().upper()
                nombre = request.POST.get('nombre', '').strip()
                descripcion = request.POST.get('descripcion', '').strip()
                categoria_id = request.POST.get('categoria')
                unidad_medida = request.POST.get('unidad_medida', 'unidad')
                stock_minimo = int(request.POST.get('stock_minimo', 0))
                maneja_lotes = request.POST.get('maneja_lotes') == 'on'
                maneja_vencimiento = request.POST.get('maneja_vencimiento') == 'on'
                activo = request.POST.get('activo') == 'on'
                
                # ✅ MANEJO DE PRECIO CON HISTORIAL
                precio_str = request.POST.get('precio', '').strip()
                precio_anterior = producto.precio
                
                print(f"💰 Precio anterior: {precio_anterior}")
                print(f"💰 Precio nuevo recibido: '{precio_str}'")
                
                if precio_str:
                    try:
                        nuevo_precio = float(precio_str)
                        if nuevo_precio >= 0:
                            # ✅ CREAR HISTORIAL SOLO SI EL PRECIO CAMBIÓ
                            if precio_anterior != nuevo_precio:
                                historial = crear_historial_precio_producto(
                                    producto=producto,
                                    precio_nuevo=nuevo_precio,
                                    usuario=request.user,
                                    observaciones='Actualización desde formulario de edición'
                                )
                                
                                if historial:
                                    print(f"✅ Historial creado: ID {historial.id}")
                                
                                # ✅ ACTUALIZAR PRECIO Y FECHA
                                producto.precio = nuevo_precio
                                producto.precio_actualizado = timezone.now()
                                print(f"✅ Precio actualizado: ${nuevo_precio}")
                            else:
                                print("ℹ️ Precio sin cambios, no se actualiza historial")
                        else:
                            messages.warning(request, 'El precio no puede ser negativo.')
                    except ValueError:
                        messages.warning(request, 'Precio inválido. Use formato: 123.45')
                elif precio_str == '':
                    # Si se deja vacío, se mantiene el precio actual
                    print("ℹ️ Precio vacío, manteniendo precio actual")
                
                # Validaciones básicas
                if not codigo or not nombre:
                    print("❌ ERROR: Código o nombre vacíos")
                    messages.error(request, 'El código y nombre son obligatorios.')
                    return render(request, 'inventario/productos/editar.html', {
                        'producto': producto,
                        'categorias': Categoria.objects.all().order_by('nombre'),
                        'etiquetas': Etiqueta.objects.all()
                    })
                
                # Verificar código único (excepto el producto actual)
                if Producto.objects.filter(codigo=codigo).exclude(id=producto.id).exists():
                    print(f"❌ ERROR: Código {codigo} ya existe")
                    messages.error(request, f'Ya existe otro producto con el código "{codigo}".')
                    return render(request, 'inventario/productos/editar.html', {
                        'producto': producto,
                        'categorias': Categoria.objects.all().order_by('nombre'),
                        'etiquetas': Etiqueta.objects.all()
                    })
                
                # Obtener categoría
                categoria = get_object_or_404(Categoria, id=categoria_id) if categoria_id else None
                print(f"📂 Categoría: {categoria}")
                
                # === MANEJO DE IMAGEN CON DEBUG DETALLADO ===
                imagen_nueva = request.FILES.get('imagen')
                eliminar_imagen = request.POST.get('eliminar_imagen') == 'on'
                
                print(f"🖼️  imagen_nueva: {imagen_nueva}")
                print(f"🗑️  eliminar_imagen: {eliminar_imagen}")
                print(f"📷 producto.imagen ANTES: '{producto.imagen}'")
                
                if eliminar_imagen and producto.imagen:
                    print("🗑️  ELIMINANDO imagen actual...")
                    if hasattr(producto, 'delete_imagen_anterior'):
                        producto.delete_imagen_anterior()
                    else:
                        if hasattr(producto.imagen, 'path') and os.path.isfile(producto.imagen.path):
                            os.remove(producto.imagen.path)
                    producto.imagen = None
                    print("✅ Imagen eliminada")
                    
                elif imagen_nueva:
                    print(f"📷 NUEVA IMAGEN detectada:")
                    print(f"   - Nombre: {imagen_nueva.name}")
                    print(f"   - Tamaño: {imagen_nueva.size} bytes")
                    print(f"   - Tipo: {imagen_nueva.content_type}")
                    
                    # Validar nueva imagen
                    ext = os.path.splitext(imagen_nueva.name)[1].lower()
                    print(f"   - Extensión: {ext}")
                    
                    if ext not in ['.jpg', '.jpeg', '.png', '.gif', '.webp']:
                        print("❌ FORMATO INVÁLIDO")
                        messages.error(request, 'Formato de imagen no válido. Use JPG, PNG, GIF o WebP.')
                        return render(request, 'inventario/productos/editar.html', {
                            'producto': producto,
                            'categorias': Categoria.objects.all().order_by('nombre'),
                            'etiquetas': Etiqueta.objects.all()
                        })
                    
                    if imagen_nueva.size > 5 * 1024 * 1024:
                        print("❌ ARCHIVO MUY GRANDE")
                        messages.error(request, 'La imagen es demasiado grande. Máximo 5MB.')
                        return render(request, 'inventario/productos/editar.html', {
                            'producto': producto,
                            'categorias': Categoria.objects.all().order_by('nombre'),
                            'etiquetas': Etiqueta.objects.all()
                        })
                    
                    # Eliminar imagen anterior si existe
                    if producto.imagen:
                        print("🗑️  Eliminando imagen anterior...")
                        if hasattr(producto, 'delete_imagen_anterior'):
                            producto.delete_imagen_anterior()
                        else:
                            if hasattr(producto.imagen, 'path') and os.path.isfile(producto.imagen.path):
                                os.remove(producto.imagen.path)
                    
                    # ASIGNAR NUEVA IMAGEN
                    print("🔄 ASIGNANDO nueva imagen...")
                    producto.imagen = imagen_nueva
                    print(f"✅ Imagen asignada: {producto.imagen}")
                else:
                    print("ℹ️  No hay cambios en la imagen")
                
                # Actualizar resto de campos
                print("📝 Actualizando campos del producto...")
                producto.codigo = codigo
                producto.nombre = nombre
                producto.descripcion = descripcion
                producto.categoria = categoria
                producto.unidad_medida = unidad_medida
                producto.stock_minimo = stock_minimo
                producto.maneja_lotes = maneja_lotes
                producto.maneja_vencimiento = maneja_vencimiento
                producto.activo = activo
                
                # GUARDAR PRODUCTO
                print("💾 GUARDANDO producto...")
                producto.save()
                print(f"✅ PRODUCTO GUARDADO!")
                print(f"📷 producto.imagen DESPUÉS: '{producto.imagen}'")
                
                if producto.imagen:
                    print(f"📁 Ruta: {producto.imagen.path}")
                    print(f"🔗 URL: {producto.imagen.url}")
                    print(f"💾 ¿Archivo existe?: {os.path.exists(producto.imagen.path)}")
                
                # Manejar etiquetas
                etiquetas_ids = request.POST.getlist('etiquetas')
                if etiquetas_ids:
                    etiquetas = Etiqueta.objects.filter(id__in=etiquetas_ids)
                    producto.etiquetas.set(etiquetas)
                else:
                    producto.etiquetas.clear()
                
                print("🎉 PROCESO COMPLETADO")
                messages.success(request, f'Producto "{producto.nombre}" actualizado exitosamente.')
                return redirect('detalle_producto', producto_id=producto.id)
                
        except Exception as e:
            print(f"❌ ERROR CRÍTICO: {str(e)}")
            import traceback
            traceback.print_exc()
            messages.error(request, f'Error al actualizar el producto: {str(e)}')
    
    # GET request o error en POST
    context = {
        'producto': producto,
        'categorias': Categoria.objects.all().order_by('nombre'),
        'etiquetas': Etiqueta.objects.all()
    }
    return render(request, 'inventario/productos/editar.html', context)

def crear_historial_precio_producto(producto, precio_nuevo, usuario=None, observaciones='', precio_anterior=None):
    from .models import HistorialPrecios
    from django.utils import timezone
    
    if precio_anterior is None:
        precio_anterior = producto.precio

    if precio_anterior != precio_nuevo:
        try:
            historial = HistorialPrecios.objects.create(
                producto=producto,
                proveedor=None,
                precio=precio_nuevo,
                precio_anterior=precio_anterior,
                usuario=usuario,
                fecha=timezone.now(),
                observaciones=observaciones,
                motivo='Actualización manual desde formulario'
            )
            print(f"✅ Historial creado exitosamente: ID {historial.id}")
            return historial
        except Exception as e:
            print(f"❌ Error al crear historial: {e}")
            return None
    else:
        print("ℹ️ Precio sin cambios, no se crea historial")
        return None

@gestor_o_admin
def actualizar_precio_producto(request, producto_id):
    """Vista específica para actualizar solo el precio de un producto"""
    producto = get_object_or_404(Producto, id=producto_id)
    
    if request.method == 'POST':
        form = PrecioProductoForm(request.POST, instance=producto)
        if form.is_valid():
            precio_anterior = producto.precio
            nuevo_precio = form.cleaned_data['precio']
            observaciones = form.cleaned_data.get('observaciones', '')
            
            # Actualizar precio
            producto.actualizar_precio(nuevo_precio)
            
            # Crear registro en historial si existe
            try:
                # Si tienes un modelo de historial de precios, crear registro aquí
                HistorialPrecios.objects.create(
                    producto=producto,
                    precio_anterior=precio_anterior,
                    precio_nuevo=nuevo_precio,
                    usuario=request.user,
                    observaciones=observaciones,
                    fecha=timezone.now()
                )
            except:
                pass  # Si no existe el modelo, continuar sin error
            
            messages.success(
                request, 
                f'Precio actualizado de ${precio_anterior or "N/A"} a ${nuevo_precio}'
            )
            return redirect('detalle_producto', producto_id=producto.id)
    else:
        form = PrecioProductoForm(instance=producto)
    
    context = {
        'producto': producto,
        'form': form,
        'precio_actual': producto.precio
    }
    return render(request, 'inventario/productos/actualizar_precio.html', context)

@todos_los_roles
def lista_movimientos(request):
    """Lista de todos los movimientos de inventario"""
    movimientos = MovimientoInventario.objects.all().select_related(
        'producto', 'sucursal', 'sucursal_destino', 'usuario'
    )
    
    # Filtros
    tipo_filter = request.GET.get('tipo')
    sucursal_filter = request.GET.get('sucursal')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    
    if tipo_filter:
        movimientos = movimientos.filter(tipo=tipo_filter)
    
    if sucursal_filter:
        movimientos = movimientos.filter(sucursal_id=sucursal_filter)
    
    if fecha_desde:
        try:
            fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            movimientos = movimientos.filter(fecha__date__gte=fecha_desde_obj)
        except ValueError:
            pass
    
    if fecha_hasta:
        try:
            fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            movimientos = movimientos.filter(fecha__date__lte=fecha_hasta_obj)
        except ValueError:
            pass
    
    # Paginación
    paginator = Paginator(movimientos, 25)
    page_number = request.GET.get('page')
    movimientos_paginados = paginator.get_page(page_number)
    
    context = {
        'movimientos': movimientos_paginados,
        'sucursales': Sucursal.objects.filter(activa=True),
        'tipos_movimiento': MovimientoInventario.TIPO_CHOICES,
        'filtros': {
            'tipo': tipo_filter,
            'sucursal': sucursal_filter,
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
        }
    }
    
    return render(request, 'inventario/movimientos/lista.html', context)

@gestor_o_admin
def actualizar_precio_ajax(request, producto_id):
    """Vista AJAX para actualizar precio rápidamente"""
    if request.method == 'POST':
        producto = get_object_or_404(Producto, id=producto_id)
        
        try:
            nuevo_precio = float(request.POST.get('precio'))
            if nuevo_precio < 0:
                return JsonResponse({'success': False, 'error': 'El precio no puede ser negativo'})
            
            precio_anterior = producto.precio
            producto.actualizar_precio(nuevo_precio)
            
            return JsonResponse({
                'success': True,
                'precio_anterior': str(precio_anterior) if precio_anterior else 'N/A',
                'precio_nuevo': str(nuevo_precio),
                'precio_formateado': producto.get_precio_formateado(),
                'mensaje': f'Precio actualizado a {producto.get_precio_formateado()}'
            })
            
        except (ValueError, TypeError):
            return JsonResponse({'success': False, 'error': 'Precio inválido'})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

@gestor_o_admin
def crear_proveedor(request):
    """Crear nuevo proveedor"""
    if request.method == 'POST':
        form = ProveedorForm(request.POST)
        if form.is_valid():
            proveedor = form.save()
            messages.success(request, f'Proveedor "{proveedor.nombre}" creado exitosamente.')
            return redirect('detalle_proveedor', proveedor_id=proveedor.id)
    else:
        form = ProveedorForm()
    
    return render(request, 'inventario/proveedores/crear.html', {'form': form})

@todos_los_roles
def detalle_proveedor(request, proveedor_id):
    """Detalle de un proveedor"""
    proveedor = get_object_or_404(Proveedor, id=proveedor_id)
    
    # Historial de precios de este proveedor
    historial_precios = HistorialPrecios.objects.filter(
        proveedor=proveedor
    ).select_related('producto')[:20]
    
    # Lotes de este proveedor
    lotes = Lote.objects.filter(
        proveedor=proveedor
    ).select_related('producto')[:20]
    
    context = {
        'proveedor': proveedor,
        'historial_precios': historial_precios,
        'lotes': lotes,
    }
    
    return render(request, 'inventario/proveedores/detalle.html', context)

@gestor_o_admin
def editar_proveedor(request, proveedor_id):
    """Editar un proveedor"""
    proveedor = get_object_or_404(Proveedor, id=proveedor_id)
    
    if request.method == 'POST':
        form = ProveedorForm(request.POST, instance=proveedor)
        if form.is_valid():
            proveedor = form.save()
            messages.success(request, f'Proveedor "{proveedor.nombre}" actualizado exitosamente.')
            return redirect('detalle_proveedor', proveedor_id=proveedor.id)
    else:
        form = ProveedorForm(instance=proveedor)
    
    return render(request, 'inventario/proveedores/editar.html', {
        'form': form, 
        'proveedor': proveedor
    })

@todos_los_roles
def lista_lotes(request, producto_id):
    """Lista de lotes para un producto"""
    producto = get_object_or_404(Producto, id=producto_id)
    
    if not producto.maneja_lotes:
        messages.warning(request, 'Este producto no maneja lotes.')
        return redirect('detalle_producto', producto_id=producto.id)
    
    lotes = Lote.objects.filter(producto=producto).select_related('proveedor').order_by('-fecha_compra')
    
    # Separar lotes por estado
    lotes_vigentes = []
    lotes_por_vencer = []
    lotes_vencidos = []
    
    for lote in lotes:
        if lote.esta_vencido:
            lotes_vencidos.append(lote)
        elif lote.dias_vencimiento and lote.dias_vencimiento <= 30:
            lotes_por_vencer.append(lote)
        else:
            lotes_vigentes.append(lote)
    
    context = {
        'producto': producto,
        'lotes_vigentes': lotes_vigentes,
        'lotes_por_vencer': lotes_por_vencer,
        'lotes_vencidos': lotes_vencidos,
        'total_lotes': lotes.count(),
    }
    
    return render(request, 'inventario/lotes/lista.html', context)

@gestor_o_admin
def crear_lote(request, producto_id):
    """Crear nuevo lote para un producto"""
    producto = get_object_or_404(Producto, id=producto_id)
    
    if not producto.maneja_lotes:
        messages.warning(request, 'Este producto no maneja lotes.')
        return redirect('detalle_producto', producto_id=producto.id)
    
    if request.method == 'POST':
        form = LoteForm(request.POST)
        if form.is_valid():
            lote = form.save(commit=False)
            lote.producto = producto
            lote.save()
            
            # Crear entrada en historial de precios
            HistorialPrecios.objects.create(
                producto=producto,
                proveedor=lote.proveedor,
                precio=lote.precio_compra,
                usuario=request.user
            )
            
            # Si tiene fecha de vencimiento, verificar alertas
            if lote.fecha_vencimiento:
                fecha_limite = timezone.now().date() + timedelta(days=30)
                if lote.fecha_vencimiento <= fecha_limite:
                    # Crear alertas para todas las sucursales
                    for sucursal in Sucursal.objects.filter(activa=True):
                        if lote.esta_vencido:
                            tipo_alerta = 'vencido'
                            mensaje = f'¡ATENCIÓN! Lote {lote.codigo} de {producto.nombre} está vencido.'
                        else:
                            tipo_alerta = 'vencimiento'
                            mensaje = f'Lote {lote.codigo} de {producto.nombre} vence el {lote.fecha_vencimiento.strftime("%d/%m/%Y")}.'
                        
                        AlertaStock.objects.get_or_create(
                            tipo=tipo_alerta,
                            producto=producto,
                            sucursal=sucursal,
                            lote=lote,
                            defaults={'mensaje': mensaje}
                        )
            
            messages.success(request, f'Lote "{lote.codigo}" creado exitosamente.')
            return redirect('lista_lotes', producto_id=producto.id)
    else:
        form = LoteForm()
    
    context = {
        'form': form,
        'producto': producto,
    }
    
    return render(request, 'inventario/lotes/crear.html', context)

@gestor_o_admin
def editar_lote(request, lote_id):
    """Editar un lote existente"""
    lote = get_object_or_404(Lote, id=lote_id)
    
    if request.method == 'POST':
        form = LoteForm(request.POST, instance=lote)
        if form.is_valid():
            lote_actualizado = form.save()
            
            # Actualizar historial de precios si cambió el precio
            if 'precio_compra' in form.changed_data:
                HistorialPrecios.objects.create(
                    producto=lote.producto,
                    proveedor=lote.proveedor,
                    precio=lote.precio_compra,
                    usuario=request.user
                )
            
            messages.success(request, f'Lote "{lote.codigo}" actualizado exitosamente.')
            return redirect('lista_lotes', producto_id=lote.producto.id)
    else:
        form = LoteForm(instance=lote)
    
    context = {
        'form': form,
        'lote': lote,
        'producto': lote.producto,
    }
    
    return render(request, 'inventario/lotes/editar.html', context)

@todos_los_roles
def detalle_lote(request, lote_id):
    """Detalle de un lote específico"""
    lote = get_object_or_404(Lote, id=lote_id)
    
    # Movimientos relacionados con este lote
    movimientos = MovimientoInventario.objects.filter(
        lote=lote
    ).select_related('sucursal', 'usuario').order_by('-fecha')
    
    context = {
        'lote': lote,
        'producto': lote.producto,
        'movimientos': movimientos,
    }
    
    return render(request, 'inventario/lotes/detalle.html', context)

@todos_los_roles
def lista_categorias(request):
    """Lista de categorías"""
    categorias = Categoria.objects.all()
    
    context = {
        'categorias': categorias,
    }
    
    return render(request, 'inventario/categorias/lista.html', context)

@gestor_o_admin
def crear_categoria(request):
    """Crear nueva categoría"""
    if request.method == 'POST':
        form = CategoriaForm(request.POST)
        if form.is_valid():
            categoria = form.save()
            messages.success(request, f'Categoría "{categoria.nombre}" creada exitosamente.')
            return redirect('lista_categorias')
    else:
        form = CategoriaForm()
    
    return render(request, 'inventario/categorias/crear.html', {'form': form})

@todos_los_roles
def lista_etiquetas(request):
    """Lista de etiquetas"""
    etiquetas = Etiqueta.objects.all().order_by('nombre')
    
    # Contar productos por etiqueta
    etiquetas_con_conteo = []
    for etiqueta in etiquetas:
        conteo = etiqueta.producto_set.filter(activo=True).count()
        etiquetas_con_conteo.append({
            'etiqueta': etiqueta,
            'productos_count': conteo
        })
    
    context = {
        'etiquetas_con_conteo': etiquetas_con_conteo,
    }
    
    return render(request, 'inventario/etiquetas/lista.html', context)

@gestor_o_admin
def crear_etiqueta(request):
    """Crear nueva etiqueta"""
    if request.method == 'POST':
        form = EtiquetaForm(request.POST)
        if form.is_valid():
            etiqueta = form.save()
            messages.success(request, f'Etiqueta "{etiqueta.nombre}" creada exitosamente.')
            return redirect('lista_etiquetas')
    else:
        form = EtiquetaForm()
    
    return render(request, 'inventario/etiquetas/crear.html', {'form': form})

@gestor_o_admin
def editar_etiqueta(request, etiqueta_id):
    """Editar una etiqueta"""
    etiqueta = get_object_or_404(Etiqueta, id=etiqueta_id)
    
    if request.method == 'POST':
        form = EtiquetaForm(request.POST, instance=etiqueta)
        if form.is_valid():
            etiqueta = form.save()
            messages.success(request, f'Etiqueta "{etiqueta.nombre}" actualizada exitosamente.')
            return redirect('lista_etiquetas')
    else:
        form = EtiquetaForm(instance=etiqueta)
    
    context = {
        'form': form,
        'etiqueta': etiqueta,
    }
    
    return render(request, 'inventario/etiquetas/editar.html', context)

@todos_los_roles
def lista_sucursales(request):
    """Lista de sucursales"""
    sucursales = Sucursal.objects.all()
    
    # Calcular estadísticas
    total_sucursales = sucursales.count()
    sucursales_activas = sucursales.filter(activa=True).count()
    sucursales_inactivas = sucursales.filter(activa=False).count()
    
    context = {
        'sucursales': sucursales,
        # Estadísticas
        'total_sucursales': total_sucursales,
        'sucursales_activas': sucursales_activas,
        'sucursales_inactivas': sucursales_inactivas,
    }
    
    return render(request, 'inventario/sucursales/lista.html', context)

@solo_administrador
def crear_sucursal(request):
    """Crear nueva sucursal"""
    if request.method == 'POST':
        form = SucursalForm(request.POST)
        if form.is_valid():
            sucursal = form.save()
            
            # Crear inventario para todos los productos existentes
            productos = Producto.objects.filter(activo=True)
            for producto in productos:
                Inventario.objects.create(
                    producto=producto,
                    sucursal=sucursal,
                    cantidad=0
                )
            
            messages.success(request, f'Sucursal "{sucursal.nombre}" creada exitosamente.')
            return redirect('lista_sucursales')
    else:
        form = SucursalForm()
    
    return render(request, 'inventario/sucursales/crear.html', {'form': form})

@solo_administrador
def editar_sucursal(request, sucursal_id):
    """Editar una sucursal"""
    sucursal = get_object_or_404(Sucursal, id=sucursal_id)
    
    if request.method == 'POST':
        form = SucursalForm(request.POST, instance=sucursal)
        if form.is_valid():
            sucursal = form.save()
            messages.success(request, f'Sucursal "{sucursal.nombre}" actualizada exitosamente.')
            return redirect('lista_sucursales')
    else:
        form = SucursalForm(instance=sucursal)
    
    return render(request, 'inventario/sucursales/editar.html', {
        'form': form, 
        'sucursal': sucursal
    })

@todos_los_roles
def reportes(request):
    """Página de reportes"""
    return render(request, 'inventario/reportes/index.html')

@todos_los_roles
def generar_reporte(request):
    """Generar reporte específico"""
    if request.method == 'POST':
        form = ReporteForm(request.POST)
        if form.is_valid():
            tipo_reporte = form.cleaned_data['tipo']
            sucursal = form.cleaned_data.get('sucursal')
            fecha_desde = form.cleaned_data.get('fecha_desde')
            fecha_hasta = form.cleaned_data.get('fecha_hasta')
            categoria = form.cleaned_data.get('categoria')
            
            if tipo_reporte == 'stock':
                return generar_reporte_stock(request, sucursal, categoria)
            elif tipo_reporte == 'movimientos':
                return generar_reporte_movimientos(request, sucursal, fecha_desde, fecha_hasta)
            elif tipo_reporte == 'alertas':
                return generar_reporte_alertas(request, sucursal)
            elif tipo_reporte == 'vencimientos':
                return generar_reporte_vencimientos(request, sucursal)
    else:
        form = ReporteForm()
    
    return render(request, 'inventario/reportes/generar.html', {'form': form})

# API endpoints para AJAX
@todos_los_roles
def api_producto_stock(request, producto_id):
    """API endpoint para obtener stock de un producto por sucursal"""
    producto = get_object_or_404(Producto, id=producto_id)
    inventarios = Inventario.objects.filter(producto=producto).select_related('sucursal')
    
    data = {
        'producto': {
            'id': producto.id,
            'nombre': producto.nombre,
            'codigo': producto.codigo,
        },
        'inventarios': []
    }
    
    for inventario in inventarios:
        data['inventarios'].append({
            'sucursal_id': inventario.sucursal.id,
            'sucursal_nombre': inventario.sucursal.nombre,
            'cantidad': inventario.cantidad,
            'ubicacion': inventario.ubicacion,
            'necesita_reposicion': inventario.necesita_reposicion,
        })
    
    return JsonResponse(data)

@todos_los_roles
def api_buscar_productos(request):
    """API endpoint para búsqueda de productos (para autocomplete)"""
    q = request.GET.get('q', '')
    
    if len(q) < 2:
        return JsonResponse({'productos': []})
    
    productos = Producto.objects.filter(
        Q(nombre__icontains=q) | Q(codigo__icontains=q),
        activo=True
    )[:10]
    
    data = {
        'productos': [
            {
                'id': p.id,
                'codigo': p.codigo,
                'nombre': p.nombre,
                'stock_total': p.get_stock_total(),
            }
            for p in productos
        ]
    }
    
    return JsonResponse(data)

def generar_reporte_stock(request, sucursal=None, categoria=None):
    """Generar reporte de stock actual"""
    # Filtrar inventarios
    inventarios = Inventario.objects.select_related(
        'producto', 'sucursal', 'producto__categoria'
    ).filter(producto__activo=True)
    
    if sucursal:
        inventarios = inventarios.filter(sucursal=sucursal)
    if categoria:
        inventarios = inventarios.filter(producto__categoria=categoria)
    
    # Crear workbook de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Stock"
    
    # Encabezados
    headers = [
        'Código', 'Producto', 'Categoría', 'Sucursal', 
        'Stock Actual', 'Stock Mínimo', 'Estado', 'Ubicación'
    ]
    
    # Escribir encabezados con estilo
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Escribir datos
    for row, inventario in enumerate(inventarios, 2):
        ws.cell(row=row, column=1, value=inventario.producto.codigo)
        ws.cell(row=row, column=2, value=inventario.producto.nombre)
        ws.cell(row=row, column=3, value=inventario.producto.categoria.nombre if inventario.producto.categoria else 'Sin categoría')
        ws.cell(row=row, column=4, value=inventario.sucursal.nombre)
        ws.cell(row=row, column=5, value=inventario.cantidad)
        ws.cell(row=row, column=6, value=inventario.producto.stock_minimo)
        
        # Estado con color
        estado_cell = ws.cell(row=row, column=7)
        if inventario.cantidad == 0:
            estado_cell.value = "Sin stock"
            estado_cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        elif inventario.necesita_reposicion:
            estado_cell.value = "Stock bajo"
            estado_cell.fill = PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid")
        else:
            estado_cell.value = "Normal"
            estado_cell.fill = PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid")
        
        ws.cell(row=row, column=8, value=inventario.ubicacion or 'No especificada')
    
    # Ajustar anchos de columna
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    fecha_actual = timezone.now().strftime('%Y%m%d_%H%M')
    filename = f'reporte_stock_{fecha_actual}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    # Guardar workbook en respuesta
    wb.save(response)
    return response

def generar_reporte_movimientos(request, sucursal=None, fecha_desde=None, fecha_hasta=None):
    """Generar reporte de movimientos"""
    movimientos = MovimientoInventario.objects.select_related(
        'producto', 'sucursal', 'sucursal_destino', 'usuario'
    )
    
    if sucursal:
        movimientos = movimientos.filter(sucursal=sucursal)
    if fecha_desde:
        movimientos = movimientos.filter(fecha__date__gte=fecha_desde)
    if fecha_hasta:
        movimientos = movimientos.filter(fecha__date__lte=fecha_hasta)
    
    # Crear workbook de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Movimientos"
    
    # Encabezados
    headers = [
        'Fecha', 'Producto', 'Código', 'Tipo', 'Sucursal', 
        'Sucursal Destino', 'Cantidad', 'Stock Anterior', 'Stock Nuevo', 
        'Usuario', 'Motivo', 'Documento'
    ]
    
    # Escribir encabezados
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Escribir datos
    for row, movimiento in enumerate(movimientos, 2):
        ws.cell(row=row, column=1, value=movimiento.fecha.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row, column=2, value=movimiento.producto.nombre)
        ws.cell(row=row, column=3, value=movimiento.producto.codigo)
        ws.cell(row=row, column=4, value=movimiento.get_tipo_display())
        ws.cell(row=row, column=5, value=movimiento.sucursal.nombre)
        ws.cell(row=row, column=6, value=movimiento.sucursal_destino.nombre if movimiento.sucursal_destino else '')
        ws.cell(row=row, column=7, value=movimiento.cantidad)
        ws.cell(row=row, column=8, value=movimiento.cantidad_anterior)
        ws.cell(row=row, column=9, value=movimiento.cantidad_nueva)
        ws.cell(row=row, column=10, value=movimiento.usuario.get_full_name() or movimiento.usuario.username)
        ws.cell(row=row, column=11, value=movimiento.motivo or '')
        ws.cell(row=row, column=12, value=movimiento.numero_documento or '')
    
    # Ajustar anchos de columna
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    fecha_actual = timezone.now().strftime('%Y%m%d_%H%M')
    filename = f'reporte_movimientos_{fecha_actual}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response

def generar_reporte_alertas(request, sucursal=None):
    """Generar reporte de alertas activas"""
    alertas = AlertaStock.objects.filter(status='activa').select_related(
        'producto', 'sucursal', 'lote'
    )
    
    if sucursal:
        alertas = alertas.filter(sucursal=sucursal)
    
    # Crear workbook de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Alertas"
    
    # Encabezados
    headers = [
        'Tipo', 'Producto', 'Código', 'Sucursal', 'Mensaje', 
        'Fecha Creación', 'Lote'
    ]
    
    # Escribir encabezados
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Escribir datos
    for row, alerta in enumerate(alertas, 2):
        # Tipo con color
        tipo_cell = ws.cell(row=row, column=1, value=alerta.get_tipo_display())
        if alerta.tipo == 'vencido':
            tipo_cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        elif alerta.tipo == 'stock_bajo':
            tipo_cell.fill = PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid")
        else:
            tipo_cell.fill = PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid")
        
        ws.cell(row=row, column=2, value=alerta.producto.nombre)
        ws.cell(row=row, column=3, value=alerta.producto.codigo)
        ws.cell(row=row, column=4, value=alerta.sucursal.nombre)
        ws.cell(row=row, column=5, value=alerta.mensaje)
        ws.cell(row=row, column=6, value=alerta.fecha_creacion.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row, column=7, value=alerta.lote.codigo if alerta.lote else '')
    
    # Ajustar anchos de columna
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    fecha_actual = timezone.now().strftime('%Y%m%d_%H%M')
    filename = f'reporte_alertas_{fecha_actual}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response

def generar_reporte_vencimientos(request, sucursal=None):
    """Generar reporte de productos por vencer"""
    fecha_limite = timezone.now().date() + timedelta(days=30)
    lotes = Lote.objects.filter(
        fecha_vencimiento__lte=fecha_limite,
        fecha_vencimiento__gte=timezone.now().date()
    ).select_related('producto', 'proveedor')
    
    # Crear workbook de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos por Vencer"
    
    # Encabezados
    headers = [
        'Producto', 'Código', 'Lote', 'Fecha Vencimiento', 
        'Días Restantes', 'Proveedor', 'Precio Compra'
    ]
    
    # Escribir encabezados
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Escribir datos
    for row, lote in enumerate(lotes, 2):
        ws.cell(row=row, column=1, value=lote.producto.nombre)
        ws.cell(row=row, column=2, value=lote.producto.codigo)
        ws.cell(row=row, column=3, value=lote.codigo)
        ws.cell(row=row, column=4, value=lote.fecha_vencimiento.strftime('%d/%m/%Y'))
        
        # Días restantes con color
        dias_cell = ws.cell(row=row, column=5, value=lote.dias_vencimiento)
        if lote.dias_vencimiento <= 7:
            dias_cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        elif lote.dias_vencimiento <= 15:
            dias_cell.fill = PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid")
        
        ws.cell(row=row, column=6, value=lote.proveedor.nombre if lote.proveedor else '')
        ws.cell(row=row, column=7, value=float(lote.precio_compra))
    
    # Ajustar anchos de columna
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    fecha_actual = timezone.now().strftime('%Y%m%d_%H%M')
    filename = f'reporte_vencimientos_{fecha_actual}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response

@todos_los_roles
def lista_ubicaciones(request, sucursal_id):
    """Lista ubicaciones de una sucursal específica"""
    sucursal = get_object_or_404(Sucursal, id=sucursal_id)
    
    # Actualizar sesión con la sucursal actual
    request.session['sucursal_id'] = sucursal_id
    request.session['sucursal_nombre'] = sucursal.nombre
    
    ubicaciones = Ubicacion.objects.filter(sucursal=sucursal)
    
    # Calcular estadísticas
    total_ubicaciones = ubicaciones.count()
    ubicaciones_activas = ubicaciones.filter(activa=True).count()
    ubicaciones_llenas = sum(1 for u in ubicaciones if hasattr(u, 'esta_llena') and u.esta_llena)
    productos_ubicados = sum(u.get_productos_count() for u in ubicaciones if hasattr(u, 'get_productos_count'))
    
    context = {
        'ubicaciones': ubicaciones,
        'sucursal_actual': sucursal,
        'total_ubicaciones': total_ubicaciones,
        'ubicaciones_activas': ubicaciones_activas,
        'ubicaciones_llenas': ubicaciones_llenas,
        'productos_ubicados': productos_ubicados,
    }
    return render(request, 'inventario/ubicaciones/lista.html', context)

@puede_asignar_ubicaciones
def crear_ubicacion(request, sucursal_id):
    """Crear nueva ubicación en una sucursal"""
    sucursal = get_object_or_404(Sucursal, id=sucursal_id)
    
    if request.method == 'POST':
        form = UbicacionForm(request.POST, sucursal=sucursal)
        if form.is_valid():
            ubicacion = form.save(commit=False)
            ubicacion.sucursal = sucursal
            ubicacion.save()
            
            messages.success(request, f'Ubicación "{ubicacion.codigo}" creada exitosamente.')
            # CAMBIAR ESTA LÍNEA - usar sucursal_id directamente
            return redirect('lista_ubicaciones', sucursal_id=sucursal_id)
    else:
        form = UbicacionForm(sucursal=sucursal)
    
    context = {
        'form': form,
        'sucursal': sucursal,
        'sucursal_actual': sucursal,  # Para el template
    }
    
    return render(request, 'inventario/ubicaciones/crear.html', context)

@puede_asignar_ubicaciones
def editar_ubicacion(request, sucursal_id, ubicacion_id):
    """Editar ubicación existente"""
    sucursal = get_object_or_404(Sucursal, id=sucursal_id)
    ubicacion = get_object_or_404(Ubicacion, id=ubicacion_id, sucursal=sucursal)
    
    if request.method == 'POST':
        form = UbicacionForm(request.POST, instance=ubicacion, sucursal=sucursal)
        if form.is_valid():
            ubicacion = form.save()
            messages.success(request, f'Ubicación "{ubicacion.codigo}" actualizada exitosamente.')
            return redirect('detalle_ubicacion', sucursal_id=sucursal_id, ubicacion_id=ubicacion.id)
    else:
        form = UbicacionForm(instance=ubicacion, sucursal=sucursal)
    
    context = {
        'form': form,
        'ubicacion': ubicacion,
        'sucursal': sucursal,
        'sucursal_actual': sucursal,
    }
    
    return render(request, 'inventario/ubicaciones/editar.html', context)

@todos_los_roles
def detalle_ubicacion(request, sucursal_id, ubicacion_id):  # Agregar sucursal_id
    sucursal = get_object_or_404(Sucursal, id=sucursal_id)
    ubicacion = get_object_or_404(Ubicacion, id=ubicacion_id, sucursal=sucursal)
    
    # Productos en esta ubicación
    inventarios = Inventario.objects.filter(
        ubicacion_detallada=ubicacion
    ).select_related('producto', 'producto__categoria').order_by('producto__nombre')
    
    # Estadísticas
    total_productos = inventarios.count()
    total_cantidad = sum(inv.cantidad for inv in inventarios)
    productos_bajo_stock = sum(1 for inv in inventarios if inv.necesita_reposicion)
    
    context = {
        'sucursal_actual': sucursal,
        'ubicacion': ubicacion,
        'inventarios': inventarios,
        'total_productos': total_productos,
        'total_cantidad': total_cantidad,
        'productos_bajo_stock': productos_bajo_stock,
    }
    
    return render(request, 'inventario/ubicaciones/detalle.html', context)

@puede_asignar_ubicaciones
def asignar_ubicacion_producto(request, inventario_id):
    """Asignar o cambiar ubicación de un inventario específico"""
    inventario = get_object_or_404(Inventario, id=inventario_id)
    
    if request.method == 'POST':
        form = AsignarUbicacionForm(request.POST, instance=inventario, sucursal=inventario.sucursal)
        if form.is_valid():
            inventario_actualizado = form.save()
            
            # Registrar el cambio como movimiento interno
            MovimientoInventario.objects.create(
                producto=inventario.producto,
                sucursal=inventario.sucursal,
                tipo='ajuste',
                cantidad=inventario.cantidad,
                cantidad_anterior=inventario.cantidad,
                cantidad_nueva=inventario.cantidad,
                motivo=f'Cambio de ubicación a {inventario_actualizado.ubicacion_detallada or "ubicación general"}',
                usuario=request.user
            )
            
            messages.success(request, 'Ubicación actualizada exitosamente.')
            return redirect('detalle_producto', producto_id=inventario.producto.id)
    else:
        form = AsignarUbicacionForm(instance=inventario, sucursal=inventario.sucursal)
    
    context = {
        'form': form,
        'inventario': inventario,
        'producto': inventario.producto,
        'sucursal': inventario.sucursal,
        'sucursal_actual': inventario.sucursal,
    }
    
    return render(request, 'inventario/ubicaciones/asignar_producto.html', context)

@verificar_permiso_ajax(['administrador', 'gestor_inventario'])
def asignar_ubicacion(request, sucursal_id, ubicacion_id):
    sucursal = get_object_or_404(Sucursal, id=sucursal_id)
    ubicacion = get_object_or_404(Ubicacion, id=ubicacion_id, sucursal=sucursal)
    
    if request.method == 'POST':
        try:
            import json
            
            
            
            # Parsear JSON
            data = json.loads(request.body)

            
            productos_data = data.get('productos', [])

            
            if not productos_data:
                
                return JsonResponse({
                    'success': False, 
                    'error': 'No se recibieron productos para asignar',
                    'asignados': 0
                })
            
            asignados = 0
            errores = []
            
            for i, producto_data in enumerate(productos_data):
                
                
                if not isinstance(producto_data, dict):
                    error = f"Producto {i+1}: formato inválido"
                    
                    errores.append(error)
                    continue
                
                producto_id = producto_data.get('producto_id')
                cantidad = producto_data.get('cantidad', 1)
                

                
                if not producto_id:
                    error = f"Producto {i+1}: ID faltante"

                    errores.append(error)
                    continue
                
                try:

                    producto_id = int(producto_id)

                    

                    producto = Producto.objects.get(id=producto_id)

                    

                    inventario, created = Inventario.objects.get_or_create(
                        producto=producto,
                        sucursal=sucursal,
                        defaults={
                            'cantidad': 0,
                            'ubicacion': '',  
                        }
                    )
                    

                    

                    inventario.ubicacion_detallada = ubicacion
                    inventario.save()
                    

                    

                    inventario.refresh_from_db()

                    
                    asignados += 1
                    
                except Producto.DoesNotExist:
                    error = f"Producto ID {producto_id} no encontrado"

                    errores.append(error)
                except ValueError as e:
                    error = f"ID inválido {producto_id}: {e}"

                    errores.append(error)
                except Exception as e:
                    error = f"Error con producto ID {producto_id}: {str(e)}"


                    errores.append(error)
            

            if errores:
                for error in errores:
                    print(f"  - {error}")
            

            inventarios_ubicacion = Inventario.objects.filter(
                sucursal=sucursal,
                ubicacion_detallada=ubicacion
            )
            print(f"Inventarios en ubicación después: {inventarios_ubicacion.count()}")
            for inv in inventarios_ubicacion:
                print(f"  - {inv.producto.codigo}: cantidad={inv.cantidad}")
            
            print("=====================")
            
            return JsonResponse({
                'success': True,
                'asignados': asignados,
                'total_solicitados': len(productos_data),
                'errores': errores,
                'inventarios_en_ubicacion': inventarios_ubicacion.count()
            })
            
        except json.JSONDecodeError as e:
            error = f"Error JSON: {e}"

            return JsonResponse({'success': False, 'error': error, 'asignados': 0})
            
        except Exception as e:
            error = f"Error general: {e}"

            import traceback
            traceback.print_exc()
            return JsonResponse({'success': False, 'error': error, 'asignados': 0})
    

    productos_disponibles = Producto.objects.filter(activo=True)
    categorias = Categoria.objects.all()
    

    productos_con_estado = []
    for producto in productos_disponibles:
        ya_asignado = Inventario.objects.filter(
            producto=producto,
            sucursal=sucursal,
            ubicacion_detallada=ubicacion
        ).exists()
        
        stock_total = 0
        if hasattr(producto, 'get_stock_total'):
            try:
                stock_total = producto.get_stock_total()
            except:
                pass
        
        productos_con_estado.append({
            'producto': producto,
            'ya_asignado': ya_asignado,
            'stock_total': stock_total,
        })
    
    context = {
        'ubicacion': ubicacion,
        'sucursal_actual': sucursal,
        'productos_con_estado': productos_con_estado,
        'productos_disponibles': productos_disponibles,
        'categorias': categorias,
    }
    
    return render(request, 'inventario/ubicaciones/asignar.html', context)

@todos_los_roles
def mapa_ubicaciones(request, sucursal_id):
    """Mapa visual de ubicaciones de una sucursal"""
    sucursal = get_object_or_404(Sucursal, id=sucursal_id)
    ubicaciones = Ubicacion.objects.filter(sucursal=sucursal)
    

    mapa_ubicaciones = {}
    for ubicacion in ubicaciones:
        tipo = ubicacion.tipo
        if tipo not in mapa_ubicaciones:
            mapa_ubicaciones[tipo] = []
        

        productos_count = ubicacion.get_productos_count()
        cantidad_total = ubicacion.get_capacidad_usada()
        porcentaje_ocupacion = ubicacion.get_porcentaje_ocupacion()
        
        mapa_ubicaciones[tipo].append({
            'ubicacion': ubicacion,
            'productos_count': productos_count,
            'cantidad_total': cantidad_total,
            'porcentaje_ocupacion': porcentaje_ocupacion,
            'estado': 'llena' if ubicacion.esta_llena else ('atencion' if ubicacion.necesita_atencion else 'normal')
        })
    
    context = {
        'total_ubicaciones': ubicaciones.count(),
        'ubicaciones': ubicaciones,
        'sucursal_actual': sucursal,
        'sucursal': sucursal,
        'mapa_ubicaciones': mapa_ubicaciones,
        'tipos_ubicacion': Ubicacion.TIPO_CHOICES,
        'ubicaciones_activas': ubicaciones.filter(activa=True).count(),
        'ubicaciones_llenas': 0,
        'ocupacion_promedio': 50,
    }
    
    return render(request, 'inventario/ubicaciones/mapa.html', context)

@puede_asignar_ubicaciones
def eliminar_ubicacion(request, sucursal_id, ubicacion_id):
    """Eliminar ubicación"""
    sucursal = get_object_or_404(Sucursal, id=sucursal_id)
    ubicacion = get_object_or_404(Ubicacion, id=ubicacion_id, sucursal=sucursal)
    
    if request.method == 'POST':

        if ubicacion.inventario_set.exists():
            messages.warning(request, f'No se puede eliminar la ubicación "{ubicacion.codigo}" porque tiene productos asignados.')
            return redirect('detalle_ubicacion', sucursal_id=sucursal_id, ubicacion_id=ubicacion_id)
        

        codigo = ubicacion.codigo
        ubicacion.delete()
        messages.success(request, f'Ubicación "{codigo}" eliminada exitosamente.')
        return redirect('lista_ubicaciones', sucursal_id=sucursal_id)
    

    context = {
        'ubicacion': ubicacion,
        'sucursal_actual': sucursal,
    }
    return render(request, 'inventario/ubicaciones/confirmar_eliminar.html', context)

@solo_administrador
def lista_usuarios(request):
    """Lista todos los usuarios del sistema con filtros"""
    

    usuarios = User.objects.select_related('perfil').prefetch_related(
        'perfil__sucursal_asignada'
    ).order_by('-date_joined')
    

    rol_filtro = request.GET.get('rol', '')
    activo_filtro = request.GET.get('activo', '')
    sucursal_filtro = request.GET.get('sucursal', '')
    busqueda = request.GET.get('q', '')
    
    if rol_filtro:
        usuarios = usuarios.filter(perfil__rol=rol_filtro)
    
    if activo_filtro:
        activo_bool = activo_filtro == 'true'
        usuarios = usuarios.filter(perfil__activo=activo_bool)
    
    if sucursal_filtro:
        usuarios = usuarios.filter(perfil__sucursal_asignada_id=sucursal_filtro)
    
    if busqueda:
        usuarios = usuarios.filter(
            models.Q(username__icontains=busqueda) |
            models.Q(first_name__icontains=busqueda) |
            models.Q(last_name__icontains=busqueda) |
            models.Q(email__icontains=busqueda)
        )
    
    # Paginación
    paginator = Paginator(usuarios, 20)
    page_number = request.GET.get('page')
    usuarios_page = paginator.get_page(page_number)
    

    stats = {
        'total_usuarios': User.objects.count(),
        'administradores': PerfilUsuario.objects.filter(rol='administrador').count(),
        'gestores': PerfilUsuario.objects.filter(rol='gestor_inventario').count(),
        'usuarios_lectura': PerfilUsuario.objects.filter(rol='usuario_lectura').count(),
        'activos': PerfilUsuario.objects.filter(activo=True).count(),
        'inactivos': PerfilUsuario.objects.filter(activo=False).count(),
    }
    

    sucursales = Sucursal.objects.filter(activa=True)
    
    context = {
        'usuarios': usuarios_page,
        'stats': stats,
        'sucursales': sucursales,
        'filtros': {
            'rol': rol_filtro,
            'activo': activo_filtro,
            'sucursal': sucursal_filtro,
            'busqueda': busqueda,
        },
        'roles_choices': PerfilUsuario.ROLES,
    }
    
    return render(request, 'inventario/usuarios/lista.html', context)

@solo_administrador
def crear_usuario(request):
    """Crear nuevo usuario con perfil"""
    
    if request.method == 'POST':
        form = CrearUsuarioForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():

                    user = form.save(commit=False)
                    user.set_password(form.cleaned_data['password'])
                    user.save()
                    

                    perfil = user.perfil
                    perfil.rol = form.cleaned_data['rol']
                    perfil.sucursal_asignada = form.cleaned_data['sucursal_asignada']
                    perfil.creado_por = request.user
                    perfil.save()
                    

                    asignar_usuario_a_grupo(user, perfil.rol)
                    
                    messages.success(request, f'Usuario {user.username} creado exitosamente.')
                    return redirect('lista_usuarios')
                    
            except Exception as e:
                messages.error(request, f'Error al crear usuario: {str(e)}')
        else:
            messages.error(request, 'Por favor corrija los errores en el formulario.')
    else:
        form = CrearUsuarioForm()
    
    context = {
        'form': form,
        'titulo': 'Crear Nuevo Usuario',
        'sucursales': Sucursal.objects.filter(activa=True),
    }
    
    return render(request, 'inventario/usuarios/crear.html', context)

@solo_administrador
def editar_usuario(request, user_id):
    """Editar usuario existente"""
    
    user = get_object_or_404(User, id=user_id)
    

    if (user == request.user and 
        user.perfil.rol == 'administrador' and 
        PerfilUsuario.objects.filter(rol='administrador', activo=True).count() == 1):
        messages.error(request, 'No puede modificar su propio perfil siendo el único administrador activo.')
        return redirect('lista_usuarios')
    
    if request.method == 'POST':
        form = EditarUsuarioForm(request.POST, instance=user)
        if form.is_valid():
            try:
                with transaction.atomic():

                    user = form.save()
                    

                    perfil = user.perfil
                    rol_anterior = perfil.rol
                    perfil.rol = form.cleaned_data['rol']
                    perfil.sucursal_asignada = form.cleaned_data['sucursal_asignada']
                    perfil.activo = form.cleaned_data['activo']
                    perfil.save()
                    

                    if rol_anterior != perfil.rol:
                        user.groups.clear()
                        asignar_usuario_a_grupo(user, perfil.rol)
                    

                    new_password = form.cleaned_data.get('new_password')
                    if new_password:
                        user.set_password(new_password)
                        user.save()
                    
                    messages.success(request, f'Usuario {user.username} actualizado exitosamente.')
                    return redirect('lista_usuarios')
                    
            except Exception as e:
                messages.error(request, f'Error al actualizar usuario: {str(e)}')
        else:
            messages.error(request, 'Por favor corrija los errores en el formulario.')
    else:

        initial_data = {
            'rol': user.perfil.rol,
            'sucursal_asignada': user.perfil.sucursal_asignada,
            'activo': user.perfil.activo,
        }
        form = EditarUsuarioForm(instance=user, initial=initial_data)
    
    context = {
        'form': form,
        'user_editado': user,
        'titulo': f'Editar Usuario: {user.get_full_name() or user.username}',
        'sucursales': Sucursal.objects.filter(activa=True),
    }
    
    return render(request, 'inventario/usuarios/editar.html', context)

@solo_administrador
def detalle_usuario(request, user_id):
    """Ver detalles completos de un usuario"""
    
    user = get_object_or_404(User, id=user_id)
    

    movimientos_recientes = MovimientoInventario.objects.filter(
        usuario=user
    ).select_related('producto', 'sucursal').order_by('-fecha')[:10]
    

    stats = {
        'total_movimientos': MovimientoInventario.objects.filter(usuario=user).count(),
        'productos_creados': Producto.objects.filter(created_at__gte=user.date_joined).count() if user.perfil.puede_gestionar_productos() else 0,
        'alertas_resueltas': AlertaStock.objects.filter(usuario_resolucion=user).count(),
        'ultimo_acceso': user.last_login,
    }
    
    context = {
        'user_detalle': user,
        'movimientos_recientes': movimientos_recientes,
        'stats': stats,
    }
    
    return render(request, 'inventario/usuarios/detalle.html', context)

@solo_administrador
def toggle_usuario_activo(request, user_id):
    """Activar/desactivar usuario via AJAX"""
    
    if request.method == 'POST':
        user = get_object_or_404(User, id=user_id)
        

        if (user.perfil.rol == 'administrador' and 
            user.perfil.activo and
            PerfilUsuario.objects.filter(rol='administrador', activo=True).count() == 1):
            return JsonResponse({
                'success': False,
                'error': 'No se puede desactivar el único administrador activo.'
            })
        

        user.perfil.activo = not user.perfil.activo
        user.perfil.save()
        
        estado = 'activado' if user.perfil.activo else 'desactivado'
        
        return JsonResponse({
            'success': True,
            'message': f'Usuario {user.username} {estado} exitosamente.',
            'nuevo_estado': user.perfil.activo
        })
    
    return JsonResponse({'success': False, 'error': 'Método no permitido.'})

def asignar_usuario_a_grupo(user, rol):
    """Función helper para asignar usuario al grupo correspondiente"""
    from django.contrib.auth.models import Group
    

    user.groups.clear()
    

    group_map = {
        'administrador': 'Administradores',
        'gestor_inventario': 'Gestores de Inventario',
        'usuario_lectura': 'Usuarios de Lectura',
    }
    
    group_name = group_map.get(rol)
    if group_name:
        try:
            group = Group.objects.get(name=group_name)
            user.groups.add(group)
        except Group.DoesNotExist:
            print(f"Grupo {group_name} no existe. Ejecute crear_grupos_y_permisos()")

@todos_los_roles
def mi_perfil(request):
    """Vista para que cada usuario vea su propio perfil"""
    
    user = request.user
    

    stats = {
        'movimientos_realizados': MovimientoInventario.objects.filter(usuario=user).count(),
        'alertas_resueltas': AlertaStock.objects.filter(usuario_resolucion=user).count(),
        'fecha_registro': user.date_joined,
        'ultimo_acceso': user.last_login,
    }
    

    actividad_reciente = MovimientoInventario.objects.filter(
        usuario=user
    ).select_related('producto', 'sucursal').order_by('-fecha')[:5]
    
    context = {
        'stats': stats,
        'actividad_reciente': actividad_reciente,
    }
    
    return render(request, 'inventario/usuarios/mi_perfil.html', context)



def permisos_context(request):
    """Context processor para agregar permisos del usuario a todos los templates"""
    if request.user.is_authenticated and hasattr(request.user, 'perfil'):
        return {
            'permisos_usuario': obtener_permisos_usuario(request.user)
        }
    return {
        'permisos_usuario': {
            'puede_gestionar_productos': False,
            'puede_gestionar_inventario': False,
            'puede_asignar_ubicaciones': False,
            'puede_crear_usuarios': False,
            'puede_ver_alertas': False,
            'puede_descargar_reportes': False,
            'solo_lectura': True,
            'rol': 'No autenticado'
        }
    }

def sucursales_context(request):
    """Context processor para sucursales"""
    context = {}
    if request.user.is_authenticated:
        context['sucursales'] = Sucursal.objects.filter(activa=True)
    return context


@puede_ver_alertas
def api_alertas_count(request):
    """API para obtener contador de alertas activas"""
    sucursal_id = request.session.get('sucursal_id')
    
    if sucursal_id:
        count = AlertaStock.objects.filter(
            status='activa',
            sucursal_id=sucursal_id
        ).count()
    else:
        count = AlertaStock.objects.filter(status='activa').count()
    
    return JsonResponse({'count': count})

@require_POST
@login_required
def cambiar_sucursal(request):
    """Vista AJAX para cambiar sucursal activa"""
    sucursal_id = request.POST.get('sucursal_id')
    
    try:
        sucursal = Sucursal.objects.get(id=sucursal_id, activa=True)
        request.session['sucursal_id'] = sucursal.id
        request.session['sucursal_nombre'] = sucursal.nombre
        
        return JsonResponse({
            'success': True,
            'message': f'Sucursal cambiada a {sucursal.nombre}'
        })
    except Sucursal.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Sucursal no válida'
        })

# Redirección para ubicaciones
@login_required
def ubicaciones_redirect(request):
    """Redireccionar a ubicaciones con sucursal activa"""
    sucursal_id = request.session.get('sucursal_id')
    if sucursal_id:
        return redirect('lista_ubicaciones', sucursal_id=sucursal_id)
    else:
        # Buscar primera sucursal disponible
        sucursal = Sucursal.objects.filter(activa=True).first()
        if sucursal:
            request.session['sucursal_id'] = sucursal.id
            request.session['sucursal_nombre'] = sucursal.nombre
            return redirect('lista_ubicaciones', sucursal_id=sucursal.id)
        else:
            messages.error(request, 'No hay sucursales activas disponibles.')
            return redirect('dashboard')

@todos_los_roles
def cambiar_password(request):
    """Cambiar contraseña personal"""
    from .forms import CambiarPasswordForm
    
    if request.method == 'POST':
        form = CambiarPasswordForm(request.user, request.POST)
        if form.is_valid():
            try:
                request.user.set_password(form.cleaned_data['new_password'])
                request.user.save()
                
                # Actualizar sesión para evitar logout
                from django.contrib.auth import update_session_auth_hash
                update_session_auth_hash(request, request.user)
                
                messages.success(request, 'Su contraseña ha sido actualizada exitosamente.')
                return redirect('mi_perfil')
                
            except Exception as e:
                messages.error(request, f'Error al cambiar contraseña: {str(e)}')
        else:
            messages.error(request, 'Por favor corrija los errores en el formulario.')
    else:
        form = CambiarPasswordForm(request.user)
    
    context = {
        'form': form,
    }
    
    return render(request, 'inventario/usuarios/cambiar_password.html', context)

